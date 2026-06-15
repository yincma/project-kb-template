from __future__ import annotations

from pathlib import Path
from typing import Any

from kb.parsers.registry import ParsedDocument, ParsedSection


def parse_xlsx_file(path: Path, config: Any | None = None) -> ParsedDocument:
    warnings: list[str] = []
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return ParsedDocument(path=path, warnings=[f"Failed to parse {path}: openpyxl is unavailable: {exc}"])

    office_cfg = _cfg_value(_cfg_value(config, "parsing", None), "office", None)
    max_rows = int(_cfg_value(office_cfg, "max_sheet_rows", 20000) or 20000)
    rows_per_section = max(1, int(_cfg_value(office_cfg, "xlsx_rows_per_section", 500) or 500))

    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
        cached_workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return ParsedDocument(path=path, warnings=[f"Failed to parse {path}: {exc}"])

    sections: list[ParsedSection] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            cached_sheet = cached_workbook[sheet_name]
            max_row = min(sheet.max_row or 0, max_rows)
            if (sheet.max_row or 0) > max_rows:
                warnings.append(f"Skipped rows after {max_rows} in {path} sheet {sheet_name}")
            if max_row <= 0 or (sheet.max_column or 0) <= 0:
                continue

            rows: list[tuple[int, str, int]] = []
            for row_index, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=max_row, max_col=sheet.max_column),
                start=1,
            ):
                rendered_cells: list[str] = []
                for col_index, cell in enumerate(row, start=1):
                    cached_cell = cached_sheet.cell(row=row_index, column=col_index)
                    rendered_cells.append(_render_cell(cell.value, cached_cell.value))
                while rendered_cells and not rendered_cells[-1].strip():
                    rendered_cells.pop()
                if any(value.strip() for value in rendered_cells):
                    rows.append((row_index, "\t".join(rendered_cells), len(rendered_cells)))

            if not rows:
                continue

            section_count = (len(rows) + rows_per_section - 1) // rows_per_section
            for start in range(0, len(rows), rows_per_section):
                window = rows[start : start + rows_per_section]
                first_row = window[0][0]
                last_row = window[-1][0]
                max_col = max(entry[2] for entry in window)
                row_range = f"{first_row}:{last_row}"
                cell_range = f"A{first_row}:{get_column_letter(max_col)}{last_row}"
                heading = sheet_name if section_count == 1 else f"{sheet_name} rows {first_row}-{last_row}"
                sections.append(
                    ParsedSection(
                        text="\n".join(entry[1] for entry in window),
                        heading=heading,
                        sheet_name=sheet_name,
                        row_range=row_range,
                        cell_range=cell_range,
                        parser_name="xlsx",
                        source_format=".xlsx",
                        extraction_method="table",
                        asset_type="sheet",
                    )
                )
    finally:
        workbook.close()
        cached_workbook.close()

    return ParsedDocument(path=path, sections=sections, warnings=warnings)


def _render_cell(value: Any, cached_value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    if text.startswith("="):
        cached = "" if cached_value is None else str(cached_value)
        return f"{text} (cached: {cached})" if cached else text
    return text


def _cfg_value(config: Any | None, name: str, default: Any = None) -> Any:
    if config is None:
        return default
    if isinstance(config, dict):
        return config.get(name, default)
    return getattr(config, name, default)
